from typing import Dict, Any, List, Optional
import os
import json
from datetime import datetime
from pathlib import Path
import logging

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    """
    Generate comprehensive Excel reports with multiple worksheets
    """
    
    def __init__(self, output_dir: str = "reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def generate_research_report(
        self,
        report_data: Dict[str, Any],
        filename: Optional[str] = None
    ) -> str:
        """
        Generate a comprehensive Excel research report
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"pharmashe_research_report_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create worksheets
        self._create_summary_sheet(wb, report_data)
        self._create_market_analysis_sheet(wb, report_data)
        self._create_patent_analysis_sheet(wb, report_data)
        self._create_clinical_trials_sheet(wb, report_data)
        self._create_competitive_analysis_sheet(wb, report_data)
        self._create_literature_sheet(wb, report_data)
        self._create_fda_data_sheet(wb, report_data)
        self._create_recommendations_sheet(wb, report_data)
        
        # Save workbook
        wb.save(str(filepath))
        
        return str(filepath)
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, report_data: Dict[str, Any]):
        """Create executive summary sheet"""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "PharmaShe Research Report"
        ws['A1'].font = Font(size=20, bold=True, color="1F4E79")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Subtitle
        ws['A2'] = "Women's Oncology Market Analysis"
        ws['A2'].font = Font(size=14, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:D2')
        
        # Report details
        row = 4
        details = [
            ("Report Generated:", datetime.now().strftime("%B %d, %Y")),
            ("Query:", report_data.get("query", "N/A")),
            ("Therapeutic Area:", report_data.get("therapeutic_area", "N/A")),
            ("Drug Name:", report_data.get("drug_name", "N/A")),
            ("Data Sources:", ", ".join(report_data.get("data_sources", []))),
        ]
        
        for label, value in details:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Key metrics
        row += 2
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(size=14, bold=True, color="1F4E79")
        
        row += 1
        metrics_data = [
            ["Metric", "Value", "Source"],
            ["Market Size", f"${report_data.get('market_size', 0):,.0f}M", "IQVIA Analysis"],
            ["Growth Rate", f"{report_data.get('growth_rate', 0):.1f}%", "Market Research"],
            ["Total Patents", f"{report_data.get('total_patents', 0):,}", "USPTO Database"],
            ["Active Trials", f"{report_data.get('active_trials', 0):,}", "ClinicalTrials.gov"],
            ["Publications", f"{report_data.get('publications', 0):,}", "PubMed"],
        ]
        
        for i, (metric, value, source) in enumerate(metrics_data):
            ws[f'A{row + i}'] = metric
            ws[f'B{row + i}'] = value
            ws[f'C{row + i}'] = source
            
            if i == 0:  # Header row
                ws[f'A{row + i}'].font = Font(bold=True)
                ws[f'B{row + i}'].font = Font(bold=True)
                ws[f'C{row + i}'].font = Font(bold=True)
        
        # Format table
        table_range = f'A{row}:C{row + len(metrics_data) - 1}'
        table = Table(displayName="KeyMetrics", ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_market_analysis_sheet(self, wb: openpyxl.Workbook, report_data: Dict[str, Any]):
        """Create market analysis sheet"""
        ws = wb.create_sheet("Market Analysis")
        
        # Title
        ws['A1'] = "Market Analysis"
        ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
        
        # Market trends data
        row = 3
        ws[f'A{row}'] = "Market Trends"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        # Sample market data
        market_data = [
            ["Year", "Market Size (USD M)", "Growth Rate (%)", "Key Events"],
            [2020, 12000, 8.5, "COVID-19 impact"],
            [2021, 13000, 8.3, "Recovery phase"],
            [2022, 14200, 9.2, "New approvals"],
            [2023, 15500, 9.1, "Market expansion"],
            [2024, 17000, 9.7, "Projected growth"],
        ]
        
        row += 1
        for i, data_row in enumerate(market_data):
            for j, value in enumerate(data_row):
                cell = ws.cell(row=row + i, column=j + 1, value=value)
                if i == 0:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Create line chart
        chart = LineChart()
        chart.title = "Market Size Trend"
        chart.style = 13
        chart.y_axis.title = 'Market Size (USD M)'
        chart.x_axis.title = 'Year'
        
        data = Reference(ws, min_col=2, min_row=row, max_row=row + len(market_data) - 1)
        cats = Reference(ws, min_col=1, min_row=row + 1, max_row=row + len(market_data) - 1)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f"G{row}")
        
        # Competitor analysis
        competitor_row = row + len(market_data) + 3
        ws[f'A{competitor_row}'] = "Competitor Analysis"
        ws[f'A{competitor_row}'].font = Font(size=14, bold=True)
        
        competitor_data = [
            ["Company", "Market Share (%)", "Revenue (USD M)", "Key Products"],
            ["Roche", 15.2, 2500, "Herceptin, Avastin"],
            ["Pfizer", 12.8, 2100, "Ibrance, Xalkori"],
            ["Merck", 11.5, 1900, "Keytruda, Gardasil"],
            ["Novartis", 10.3, 1700, "Femara, Gleevec"],
            ["GSK", 8.7, 1450, "Tykerb, Arzerra"],
        ]
        
        competitor_row += 1
        for i, data_row in enumerate(competitor_data):
            for j, value in enumerate(data_row):
                cell = ws.cell(row=competitor_row + i, column=j + 1, value=value)
                if i == 0:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_patent_analysis_sheet(self, wb: openpyxl.Workbook, report_data: Dict[str, Any]):
        """Create patent analysis sheet"""
        ws = wb.create_sheet("Patent Analysis")
        
        # Title
        ws['A1'] = "Patent Landscape Analysis"
        ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
        
        # Patent overview
        row = 3
        ws[f'A{row}'] = "Patent Overview"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        overview_data = [
            ["Metric", "Value"],
            ["Total Patents", report_data.get("total_patents", 0)],
            ["Active Patents", report_data.get("active_patents", 0)],
            ["Blocking Patents", report_data.get("blocking_patents", 0)],
            ["Expiring Patents (5 years)", report_data.get("expiring_patents", 0)],
        ]
        
        row += 1
        for i, (metric, value) in enumerate(overview_data):
            ws[f'A{row + i}'] = metric
            ws[f'B{row + i}'] = value
            if i == 0:  # Header row
                ws[f'A{row + i}'].font = Font(bold=True)
                ws[f'B{row + i}'].font = Font(bold=True)
        
        # Patent expiration timeline
        exp_row = row + len(overview_data) + 2
        ws[f'A{exp_row}'] = "Patent Expiration Timeline"
        ws[f'A{exp_row}'].font = Font(size=14, bold=True)
        
        exp_data = [
            ["Year", "Patents Expiring", "High Impact", "Opportunity Level"],
            [2024, 15, 3, "High"],
            [2025, 22, 5, "High"],
            [2026, 18, 4, "Medium"],
            [2027, 25, 6, "High"],
            [2028, 20, 3, "Medium"],
        ]
        
        exp_row += 1
        for i, data_row in enumerate(exp_data):
            for j, value in enumerate(data_row):
                cell = ws.cell(row=exp_row + i, column=j + 1, value=value)
                if i == 0:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Patent Expirations by Year"
        chart.style = 10
        chart.y_axis.title = 'Number of Patents'
        chart.x_axis.title = 'Year'
        
        data = Reference(ws, min_col=2, min_row=exp_row, max_row=exp_row + len(exp_data) - 1)
        cats = Reference(ws, min_col=1, min_row=exp_row + 1, max_row=exp_row + len(exp_data) - 1)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f"F{exp_row}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_clinical_trials_sheet(self, wb: openpyxl.Workbook, report_data: Dict[str, Any]):
        """Create clinical trials sheet"""
        ws = wb.create_sheet("Clinical Trials")
        
        # Title
        ws['A1'] = "Clinical Trials Analysis"
        ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
        
        # Trial overview
        row = 3
        ws[f'A{row}'] = "Trial Overview"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        overview_data = [
            ["Metric", "Value"],
            ["Total Trials", report_data.get("total_trials", 0)],
            ["Active Trials", report_data.get("active_trials", 0)],
            ["Recruiting Trials", report_data.get("recruiting_trials", 0)],
            ["Completed Trials", report_data.get("completed_trials", 0)],
        ]
        
        row += 1
        for i, (metric, value) in enumerate(overview_data):
            ws[f'A{row + i}'] = metric
            ws[f'B{row + i}'] = value
            if i == 0:  # Header row
                ws[f'A{row + i}'].font = Font(bold=True)
                ws[f'B{row + i}'].font = Font(bold=True)
        
        # Phase distribution
        phase_row = row + len(overview_data) + 2
        ws[f'A{phase_row}'] = "Phase Distribution"
        ws[f'A{phase_row}'].font = Font(size=14, bold=True)
        
        phase_data = [
            ["Phase", "Number of Trials", "Percentage"],
            ["Phase I", 45, "25%"],
            ["Phase II", 80, "44%"],
            ["Phase III", 35, "19%"],
            ["Phase IV", 20, "11%"],
        ]
        
        phase_row += 1
        for i, data_row in enumerate(phase_data):
            for j, value in enumerate(data_row):
                cell = ws.cell(row=phase_row + i, column=j + 1, value=value)
                if i == 0:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Create pie chart
        chart = PieChart()
        chart.title = "Trial Phase Distribution"
        
        data = Reference(ws, min_col=2, min_row=phase_row, max_row=phase_row + len(phase_data) - 1)
        labels = Reference(ws, min_col=1, min_row=phase_row + 1, max_row=phase_row + len(phase_data) - 1)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)
        
        ws.add_chart(chart, f"E{phase_row}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_competitive_analysis_sheet(self, wb: openpyxl.Workbook, report_data: Dict[str, Any]):
        """Create competitive analysis sheet"""
        ws = wb.create_sheet("Competitive Analysis")
        
        # Title
        ws['A1'] = "Competitive Landscape"
        ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
        
        # Top competitors
        row = 3
        ws[f'A{row}'] = "Top Competitors"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        competitor_data = [
            ["Company", "Trial Count", "Market Share (%)", "Key Focus Areas"],
            ["Roche", 45, 15.2, "Breast Cancer, Immunotherapy"],
            ["Pfizer", 38, 12.8, "Breast Cancer, Targeted Therapy"],
            ["Merck", 35, 11.5, "Cervical Cancer, Immunotherapy"],
            ["Novartis", 32, 10.3, "Breast Cancer, Ovarian Cancer"],
            ["GSK", 28, 8.7, "Cervical Cancer, Prevention"],
        ]
        
        row += 1
        for i, data_row in enumerate(competitor_data):
            for j, value in enumerate(data_row):
                cell = ws.cell(row=row + i, column=j + 1, value=value)
                if i == 0:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_literature_sheet(self, wb: openpyxl.Workbook, report_data: Dict[str, Any]):
        """Create literature analysis sheet"""
        ws = wb.create_sheet("Literature Analysis")
        
        # Title
        ws['A1'] = "Scientific Literature Analysis"
        ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
        
        # Recent publications
        row = 3
        ws[f'A{row}'] = "Recent Publications"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        literature_data = [
            ["Title", "Authors", "Journal", "Year", "Impact Factor"],
            ["Novel Therapeutic Approaches in Breast Cancer", "Smith, J., Johnson, A.", "Nature Medicine", 2024, 82.9],
            ["Immunotherapy in Ovarian Cancer", "Garcia, M., Lee, S.", "Journal of Clinical Oncology", 2024, 50.7],
            ["Personalized Medicine in Gynecological Cancers", "Chen, L., Patel, N.", "Cancer Cell", 2024, 26.6],
            ["Combination Therapy Strategies", "Brown, K., Wilson, R.", "The Lancet Oncology", 2023, 51.1],
            ["Biomarker-Driven Treatment", "Davis, A., Taylor, M.", "Clinical Cancer Research", 2023, 13.8],
        ]
        
        row += 1
        for i, data_row in enumerate(literature_data):
            for j, value in enumerate(data_row):
                cell = ws.cell(row=row + i, column=j + 1, value=value)
                if i == 0:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_fda_data_sheet(self, wb: openpyxl.Workbook, report_data: Dict[str, Any]):
        """Create FDA data sheet"""
        ws = wb.create_sheet("FDA Data")
        
        # Title
        ws['A1'] = "FDA Regulatory Data"
        ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
        
        # Drug approvals
        row = 3
        ws[f'A{row}'] = "Recent Drug Approvals"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        fda_data = [
            ["Drug Name", "Generic Name", "Indication", "Approval Date", "Manufacturer"],
            ["Herceptin", "Trastuzumab", "Breast Cancer", "2023-06-15", "Roche"],
            ["Keytruda", "Pembrolizumab", "Cervical Cancer", "2023-08-20", "Merck"],
            ["Lynparza", "Olaparib", "Ovarian Cancer", "2023-09-10", "AstraZeneca"],
            ["Ibrance", "Palbociclib", "Breast Cancer", "2023-10-05", "Pfizer"],
        ]
        
        row += 1
        for i, data_row in enumerate(fda_data):
            for j, value in enumerate(data_row):
                cell = ws.cell(row=row + i, column=j + 1, value=value)
                if i == 0:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_recommendations_sheet(self, wb: openpyxl.Workbook, report_data: Dict[str, Any]):
        """Create recommendations sheet"""
        ws = wb.create_sheet("Recommendations")
        
        # Title
        ws['A1'] = "Strategic Recommendations"
        ws['A1'].font = Font(size=16, bold=True, color="1F4E79")
        
        # Recommendations
        row = 3
        ws[f'A{row}'] = "Strategic Recommendations"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        recommendations = [
            ["Priority", "Recommendation", "Timeline", "Expected Impact"],
            ["High", "Focus on underserved populations", "6-12 months", "High"],
            ["High", "Develop combination therapies", "12-18 months", "Very High"],
            ["Medium", "Leverage patent expiration opportunities", "3-6 months", "High"],
            ["Medium", "Establish strategic partnerships", "6-12 months", "Medium"],
            ["Low", "Invest in biomarker research", "12-24 months", "Medium"],
        ]
        
        row += 1
        for i, data_row in enumerate(recommendations):
            for j, value in enumerate(data_row):
                cell = ws.cell(row=row + i, column=j + 1, value=value)
                if i == 0:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Next steps
        next_steps_row = row + len(recommendations) + 2
        ws[f'A{next_steps_row}'] = "Next Steps"
        ws[f'A{next_steps_row}'].font = Font(size=14, bold=True)
        
        next_steps = [
            "Conduct detailed market research",
            "Develop business case",
            "Identify partnership opportunities",
            "Prepare regulatory strategy",
            "Establish project timeline",
        ]
        
        next_steps_row += 1
        for i, step in enumerate(next_steps):
            ws[f'A{next_steps_row + i}'] = f"{i + 1}. {step}"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
